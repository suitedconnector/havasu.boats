#!/usr/bin/env python3
"""
havasu.boats — Long-form description generator
================================================
Input : Outscraper reviews export (XLSX)
Output: descriptions-generated.json with `longDescription` (400–500 words each)
        + listings-with-descriptions.json (merged into existing listings.json)

Design:
  1. Group reviews by business name
  2. Extract review themes (what customers praise/complain about)
  3. One Claude call per business: synthesize into unique, SEO-ready narrative
  4. Output structured JSON, merge with existing listings

Setup:
  pip install openpyxl anthropic --break-system-packages
  export ANTHROPIC_API_KEY=sk-ant-...
  python3 generate-listing-descriptions.py <outscraper.xlsx> <listings.json>
"""

import json
import sys
import time
from collections import defaultdict
from pathlib import Path

import openpyxl
from anthropic import Anthropic


# ── Config ──────────────────────────────────────────────────────────────────
MODEL = "claude-opus-4-8"
MIN_REVIEWS = 3  # only generate if business has >= 3 reviews
REQUEST_TIMEOUT = 60
SLEEP_BETWEEN = 1.0  # be polite to the API


def load_listings(filepath):
    """Load existing listings.json to map business names to IDs."""
    with open(filepath) as f:
        listings = json.load(f)
    name_to_id = {l["name"]: l["id"] for l in listings}
    return listings, name_to_id


def extract_reviews_from_excel(xlsx_path):
    """
    Read Outscraper Excel, group reviews by business name.
    Return: {business_name: [review_text, review_text, ...], ...}
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    
    # Get column indices from header row
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        headers[cell.value] = col_idx
    
    name_col = headers.get("name")
    review_text_col = headers.get("review_text")
    review_rating_col = headers.get("review_rating")
    
    if not (name_col and review_text_col):
        print("Error: 'name' and 'review_text' columns not found.", file=sys.stderr)
        sys.exit(1)
    
    reviews_by_business = defaultdict(list)
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        name_cell = row[name_col - 1]
        review_cell = row[review_text_col - 1]
        rating_cell = row[review_rating_col - 1] if review_rating_col else None
        
        name = name_cell.value
        review_text = review_cell.value
        rating = rating_cell.value if rating_cell else None
        
        if name and review_text:
            reviews_by_business[name].append({
                "text": review_text,
                "rating": rating,
            })
    
    wb.close()
    return reviews_by_business


def synthesize_description(business_name, reviews):
    """
    Call Claude once per business to synthesize reviews into a 400–500 word narrative.
    Returns the longDescription text.
    """
    client = Anthropic()
    
    # Deduplicate & prepare review snippet
    review_texts = [r["text"] for r in reviews if r["text"]][:10]  # use up to 10 reviews
    review_snippet = "\n\n".join([f"— {rt[:300]}" for rt in review_texts])
    
    user_msg = f"""
Business: {business_name}

Customer Reviews (sample):
{review_snippet}

Your task: Write a unique, original 400–500 word description for this boat rental business's SEO listing page. 

Requirements:
- Lead with a strong value proposition (~50 words)
- Extract 3–4 key themes from the reviews (e.g., "friendly staff", "clean boats", "punctual service")
- Write each theme as a distinct paragraph (~80–100 words each)
- Include a closing paragraph about experience/vibe
- Use active voice, clear language, no marketing jargon
- NO exact quotes from reviews; synthesize and paraphrase
- Make it feel authentic, not templated
- Suitable for a Google Business Profile / local directory listing page

Output: Plain prose, no markdown, no section headers. Just flowing paragraphs.
"""

    resp = client.messages.create(
        model=MODEL,
        max_tokens=1024,
        messages=[{"role": "user", "content": user_msg}],
    )
    
    text = "".join(b.text for b in resp.content if b.type == "text").strip()
    return text


def main():
    if len(sys.argv) < 3:
        print(f"Usage: {sys.argv[0]} <outscraper.xlsx> <listings.json>", file=sys.stderr)
        sys.exit(1)
    
    xlsx_path = sys.argv[1]
    listings_path = sys.argv[2]
    
    # Load existing listings
    listings, name_to_id = load_listings(listings_path)
    
    # Extract reviews from Excel
    print("Reading Outscraper Excel...")
    reviews_by_business = extract_reviews_from_excel(xlsx_path)
    
    # Filter: only businesses with >= MIN_REVIEWS and in our listings
    candidates = {
        name: reviews
        for name, reviews in reviews_by_business.items()
        if len(reviews) >= MIN_REVIEWS and name in name_to_id
    }
    
    print(f"Found {len(candidates)} businesses with {MIN_REVIEWS}+ reviews.")
    
    # Generate descriptions
    descriptions = {}
    for i, (business_name, reviews) in enumerate(candidates.items(), 1):
        print(f"[{i}/{len(candidates)}] {business_name} ({len(reviews)} reviews)...")
        
        long_desc = synthesize_description(business_name, reviews)
        descriptions[business_name] = {
            "id": name_to_id[business_name],
            "name": business_name,
            "longDescription": long_desc,
            "review_count": len(reviews),
        }
        
        time.sleep(SLEEP_BETWEEN)
    
    # Write descriptions-generated.json
    out_descriptions = "data/descriptions-generated.json"
    with open(out_descriptions, "w") as f:
        json.dump(descriptions, f, indent=2)
    print(f"\nWrote {out_descriptions}")
    
    # Merge into listings (for convenience)
    for listing in listings:
        name = listing["name"]
        if name in descriptions:
            listing["longDescription"] = descriptions[name]["longDescription"]
    
    out_merged = "data/listings-with-descriptions.json"
    with open(out_merged, "w") as f:
        json.dump(listings, f, indent=2)
    print(f"Wrote {out_merged} (listings.json + new descriptions)")
    
    print(f"\nDone. Generated {len(descriptions)} descriptions.")
    print("Next: review descriptions-generated.json, then merge into listings.json")


if __name__ == "__main__":
    main()
